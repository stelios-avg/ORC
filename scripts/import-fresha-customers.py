#!/usr/bin/env python3
"""Copy name / phone / email from a customer Excel into public.patients."""
from __future__ import annotations

import json
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if not path.exists():
        return env
    for line in path.read_text().splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        env[k.strip()] = v.strip().strip('"').strip("'")
    return env


def digits_only(value: object) -> str:
    return re.sub(r"\D+", "", str(value or ""))


def normalize_name(name: str) -> str:
    name = re.sub(r"\s+", " ", name)
    name = re.sub(r"\bviber\b", "", name, flags=re.I)
    return name.strip().upper()


def names_match(a: str, b: str) -> bool:
    left, right = normalize_name(a), normalize_name(b)
    if not left or not right:
        return False
    return left == right or left in right or right in left


def normalize_phone(raw: object) -> str | None:
    text = str(raw or "").strip()
    if not text:
        return None
    digits = digits_only(text)
    if not digits:
        return None
    if digits.startswith("357") and len(digits) >= 11:
        national = digits[3:]
        if len(national) == 8:
            return f"+357 {national[:2]} {national[2:]}"
        return f"+357 {national}"
    if len(digits) == 8 and digits[0] in "29":
        return f"+357 {digits[:2]} {digits[2:]}"
    return f"+{digits}"


def cell(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_xlsx(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    header = [cell(h).lower() for h in next(rows)]
    aliases = {
        "ονοματεπώνυμο": "full",
        "όνομα": "first",
        "επώνυμο": "last",
        "αριθμός κινητού": "mobile",
        "τηλέφωνο": "landline",
        "email": "email",
    }
    cols = {aliases[h]: i for i, h in enumerate(header) if h in aliases}

    def pick(row: tuple, key: str) -> object:
        i = cols.get(key)
        return "" if i is None else row[i]

    out: list[dict] = []
    seen: set[str] = set()
    for row in rows:
        full = cell(pick(row, "full")) or f"{cell(pick(row, 'first'))} {cell(pick(row, 'last'))}".strip()
        if len(full) < 2:
            continue
        phone = normalize_phone(pick(row, "mobile")) or normalize_phone(pick(row, "landline"))
        email = cell(pick(row, "email")).lower() or None
        key = f"{normalize_name(full)}|{digits_only(phone)}|{email or ''}"
        if key in seen:
            continue
        seen.add(key)
        out.append({"name": full, "phone": phone, "email": email})
    wb.close()
    return out


def sql_lit(value: object) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def write_sql(records: list[dict], out: Path) -> None:
    lines = [
        "-- Copy name, phone, email into public.patients",
        "-- SQL Editor: https://supabase.com/dashboard/project/cptghymoeisjlmgggntt/sql/new",
        "begin;",
        "",
        "create temporary table excel_import (",
        "  name text not null,",
        "  phone text,",
        "  email text",
        ");",
        "",
    ]
    for i in range(0, len(records), 80):
        chunk = records[i : i + 80]
        lines.append("insert into excel_import (name, phone, email) values")
        values = [
            f"  ({sql_lit(r['name'])}, {sql_lit(r['phone'])}, {sql_lit(r['email'])})"
            for r in chunk
        ]
        lines.append(",\n".join(values) + ";")
        lines.append("")

    lines += [
        "update public.patients p set",
        "  phone = coalesce(p.phone, i.phone),",
        "  email = coalesce(p.email, i.email)",
        "from excel_import i",
        "where (i.email is not null and lower(coalesce(p.email, '')) = lower(i.email))",
        "   or (",
        "     length(right(regexp_replace(coalesce(p.phone, ''), '\\D', '', 'g'), 8)) = 8",
        "     and right(regexp_replace(coalesce(p.phone, ''), '\\D', '', 'g'), 8)",
        "       = right(regexp_replace(coalesce(i.phone, ''), '\\D', '', 'g'), 8)",
        "     and regexp_replace(replace(upper(p.name), 'VIBER', ''), '\\s+', ' ', 'g')",
        "       = regexp_replace(replace(upper(i.name), 'VIBER', ''), '\\s+', ' ', 'g')",
        "   );",
        "",
        "insert into public.patients (name, phone, email)",
        "select i.name, i.phone, i.email",
        "from excel_import i",
        "where not exists (",
        "  select 1 from public.patients p",
        "  where (i.email is not null and lower(coalesce(p.email, '')) = lower(i.email))",
        "     or (",
        "       length(right(regexp_replace(coalesce(p.phone, ''), '\\D', '', 'g'), 8)) = 8",
        "       and right(regexp_replace(coalesce(p.phone, ''), '\\D', '', 'g'), 8)",
        "         = right(regexp_replace(coalesce(i.phone, ''), '\\D', '', 'g'), 8)",
        "       and regexp_replace(replace(upper(p.name), 'VIBER', ''), '\\s+', ' ', 'g')",
        "         = regexp_replace(replace(upper(i.name), 'VIBER', ''), '\\s+', ' ', 'g')",
        "     )",
        ");",
        "",
        "commit;",
        "",
        "select count(*) as patients from public.patients;",
    ]
    out.write_text("\n".join(lines) + "\n", encoding="utf-8")


def rest(url: str, key: str, method: str, path: str, body: object | None = None):
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    data = None if body is None else json.dumps(body).encode()
    req = urllib.request.Request(url.rstrip("/") + path, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            return resp.status, resp.read()
    except urllib.error.HTTPError as e:
        return e.code, e.read()


def main() -> int:
    sql_out = None
    args = []
    i = 1
    while i < len(sys.argv):
        if sys.argv[i] == "--sql":
            if i + 1 < len(sys.argv) and not sys.argv[i + 1].startswith("-"):
                sql_out = Path(sys.argv[i + 1]).expanduser()
                i += 2
            else:
                sql_out = Path.home() / "Downloads" / "orc_import_patients.sql"
                i += 1
        else:
            args.append(sys.argv[i])
            i += 1

    if not args:
        print("Usage: python3 scripts/import-fresha-customers.py <excel> [--sql [outfile]]")
        return 2
    xlsx = Path(args[0]).expanduser()
    if not xlsx.exists():
        print("missing excel")
        return 2

    records = parse_xlsx(xlsx)
    print(f"parsed {len(records)}")
    if sql_out:
        write_sql(records, sql_out)
        print(f"sql {sql_out}")
        return 0

    env = {**load_env(ROOT / ".env"), **os.environ}
    url = env.get("PUBLIC_SUPABASE_URL") or ""
    key = env.get("SUPABASE_SERVICE_ROLE_KEY") or env.get("SUPABASE_ACCESS_TOKEN") or ""
    if not url or not key:
        print("missing SUPABASE_SERVICE_ROLE_KEY")
        return 3

    inserted = 0
    for i in range(0, len(records), 100):
        chunk = records[i : i + 100]
        st, body = rest(url, key, "POST", "/rest/v1/patients", chunk)
        if st not in (200, 201, 204):
            print("insert_failed", st)
            return 1
        inserted += len(chunk)
        print(f"inserted {inserted}/{len(records)}")
    print(f"done inserted={inserted}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
